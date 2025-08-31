import os
import io
import time
import pandas as pd
import streamlit as st
from utils import LLMBridge, extract_questions, read_docx_text, analyze_group, detect_language

st.set_page_config(page_title="IPSOS ANALYZER", layout="wide")

# ----------------------------- STYLES -----------------------------
PRIMARY_NAVY = "#0B1E39"  # dark navy
SKY = "#D6ECFF"           # cold sky blue
st.markdown(f"""
    <style>
    .reportview-container .main .block-container{{
        padding-top: 0rem;
        padding-bottom: 2rem;
    }}
    .top-banner {{
        background: linear-gradient(180deg, {SKY} 0%, #F7FBFF 100%);
        padding: 28px 16px 8px 16px;
        text-align: center;
        border-bottom: 2px solid {PRIMARY_NAVY}22;
    }}
    .title {{
        font-size: 42px;
        font-weight: 800;
        color: {PRIMARY_NAVY};
        letter-spacing: 1px;
    }}
    .subtitle {{
        color: {PRIMARY_NAVY};
        opacity: 0.75;
        font-size: 14px;
        margin-top: 4px;
    }}
    .card {{
        background: white;
        border: 1px solid #e6eef7;
        border-radius: 16px;
        padding: 16px;
        box-shadow: 0 6px 20px rgba(11,30,57,0.06);
    }}
    .logbox textarea {{
        font-family: ui-monospace, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
        font-size: 13px !important;
    }}
    .stProgress > div > div > div > div {{
        background-color: #1DB954;
    }}
    </style>
""", unsafe_allow_html=True)

# ----------------------------- HEADER (logo + titre) -----------------------------
with st.container():
    st.markdown('<div class="top-banner">', unsafe_allow_html=True)
    cols = st.columns([1, 2, 1])
    with cols[1]:
        logo_path = "assets/ipsos_logo_placeholder.png"
        if os.path.exists(logo_path):
            st.image(logo_path, width=140)
        else:
            st.write("📌 Logo manquant (assets/ipsos_logo_placeholder.png)")
        st.markdown('<div class="title">IPSOS ANALYZER</div>', unsafe_allow_html=True)
        st.markdown('<div class="subtitle">Analyse intelligente des guides & transcriptions — sortie multilingue professionnelle</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ----------------------------- SESSION STATE -----------------------------
if "log" not in st.session_state:
    st.session_state.log = ["Bienvenue, merci d'entrer la clé OpenAI, sélectionner le guide, choisir la/les transcription(s) et cliquer sur « Analyser → Excel »."]

if "transcriptions" not in st.session_state:
    st.session_state.transcriptions = {}  # name -> bytes
if "selected_to_remove" not in st.session_state:
    st.session_state.selected_to_remove = []
if "generated_excels" not in st.session_state:
    st.session_state.generated_excels = []  # list of (name, bytes)

def log(msg: str):
    ts = time.strftime("%H:%M:%S")
    st.session_state.log.append(f"[{ts}] {msg}")

def progress_set(p: float):
    st.session_state._progress_val = max(0.0, min(1.0, p))

# ----------------------------- LAYOUT -----------------------------
c1, c2, c3 = st.columns([1.2, 1.1, 1.1])

with c1:
    st.markdown("#### 1) Clé OpenAI")
    paid_or_free = st.radio("Type de clé", ["Payante", "Gratuite (mode démo)"], horizontal=True)
    openai_key = st.text_input("Saisir la clé API (masquée)", type="password", placeholder="sk-...")
    base_url = st.text_input("Base URL (optionnel, pour proxy/compatibles)", placeholder="laisser vide sauf besoin spécifique")

    if st.button("Tester la clé"):
        try:
            bridge = LLMBridge(api_key=openai_key or "DUMMY", base_url=base_url or None, paid=(paid_or_free == "Payante"))
            out = bridge.chat(
                [{"role": "system", "content": "Tu es un assistant."},
                 {"role": "user", "content": "Dis simplement OK."}],
                model="gpt-4o", temperature=0.0
            )
            st.success("Test réussi ✅" if "OK" in (out or "").upper() else "Réponse reçue ✅")
            log("Test de clé: succès.")
        except Exception as e:
            st.error(f"Echec du test : {e}")
            log(f"Echec du test de clé : {e}")

    st.markdown("---")
    st.markdown("#### 2) Langue de sortie")
    lang = st.radio("Choisir la langue de sortie", ["Français", "Arabe dialectale (Darija)", "Anglais"], horizontal=True)

    st.markdown("---")
    st.markdown("#### 3) Guide des questions (.docx)")
    guide_file = st.file_uploader("Sélectionner le fichier .docx", type=["docx"], accept_multiple_files=False)
    colg1, colg2 = st.columns(2)
    with colg1:
        if st.button("Charger le guide"):
            if guide_file is not None:
                st.session_state["guide_bytes"] = guide_file.read()
                st.session_state["guide_name"] = guide_file.name
                log(f"Guide chargé: {guide_file.name}")
            else:
                st.warning("Veuillez sélectionner un fichier .docx.")
    with colg2:
        if st.button("Vider le guide"):
            st.session_state.pop("guide_bytes", None)
            st.session_state.pop("guide_name", None)
            log("Guide vidé.")

with c2:
    st.markdown("#### 4) Transcriptions (.docx)")
    tr_files = st.file_uploader("Ajouter une ou plusieurs transcriptions", type=["docx"], accept_multiple_files=True)
    if st.button("Ajouter au lot"):
        if tr_files:
            for f in tr_files:
                st.session_state.transcriptions[f.name] = f.read()
            log(f"{len(tr_files)} transcription(s) ajoutée(s).")
        else:
            st.warning("Aucun fichier sélectionné.")

    st.markdown("Transcriptions en lot:")
    if st.session_state.transcriptions:
        names = list(st.session_state.transcriptions.keys())
        st.session_state.selected_to_remove = st.multiselect("Sélectionner pour retirer", names, default=[])
        colr1, colr2 = st.columns(2)
        with colr1:
            if st.button("Retirer la sélection"):
                for n in st.session_state.selected_to_remove:
                    st.session_state.transcriptions.pop(n, None)
                log(f"Transcriptions retirées: {', '.join(st.session_state.selected_to_remove) or 'aucune'}")
        with colr2:
            if st.button("Vider toutes"):
                st.session_state.transcriptions.clear()
                log("Toutes les transcriptions ont été vidées.")
    else:
        st.info("Aucune transcription pour l'instant.")

    st.markdown("---")
    st.markdown("#### 5) Barre de progression")
    st.session_state._progress_val = st.session_state.get("_progress_val", 0.0)
    prog = st.progress(st.session_state._progress_val)

with c3:
    st.markdown("#### 6) Actions")
    excel_name = st.text_input("Nommer l'Excel (sans extension)", value="analyse_ipsos")

    # ---------- Excel styling helper ----------
    def build_styled_excel(df: pd.DataFrame) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()
        ws = wb.active
        ws.title = "Analyse"

        # Écrire le DF
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Styles
        header_fill = PatternFill("solid", fgColor="0B1E39")  # navy
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
        body_font = Font(name="Calibri", size=11)
        wrap = Alignment(wrap_text=True, vertical="top")
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Corps + zébrage
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = body_font
                cell.alignment = wrap
                cell.border = border
            # zébrage une ligne sur deux
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="F7FBFF")

        # Largeurs colonnes
        for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            header_text = ws.cell(row=1, column=col_idx).value or ""
            if header_text == "Question":
                width = 60
            else:
                width = 50
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # Gel des volets
        ws.freeze_panes = "A2"

        # Ligne de titre au-dessus de la table (facultatif)
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        tcell = ws.cell(row=1, column=1, value="IPSOS ANALYZER — Résultats")
        tcell.font = Font(name="Calibri", size=15, bold=True, color="0B1E39")
        tcell.alignment = Alignment(horizontal="center", vertical="center")

        buff = io.BytesIO()
        wb.save(buff)
        return buff.getvalue()

    if st.button("Analyser → Excel"):
        if "guide_bytes" not in st.session_state:
            st.error("Veuillez charger d'abord un guide des questions (.docx).")
        elif not st.session_state.transcriptions:
            st.error("Veuillez ajouter au moins une transcription (.docx).")
        else:
            # Langue cible
            target_lang = {"Français": "français",
                           "Arabe dialectale (Darija)": "arabe dialectale (darija marocaine)",
                           "Anglais": "anglais"}[lang]

            # Extraction questions
            guide_bytes = st.session_state["guide_bytes"]
            questions = extract_questions(guide_bytes)
            if not questions:
                st.error("Impossible d'extraire des questions du guide. Assurez-vous que le guide contient des questions claires.")
            else:
                # Détection langue guide (info journal)
                try:
                    guide_text = read_docx_text(guide_bytes)
                except Exception:
                    guide_text = ""
                guide_lang = detect_language(guide_text) if guide_text else "inconnue"
                log(f"{len(questions)} question(s) détectée(s) dans le guide. Langue détectée: {guide_lang}.")

                # Bridge (clé)
                st.session_state["openai_key"] = openai_key
                st.session_state["base_url"] = (base_url or None)
                bridge = LLMBridge(
                    api_key=(os.environ.get("OPENAI_API_KEY") or openai_key or "DUMMY"),
                    base_url=(base_url or None),
                    paid=(paid_or_free == "Payante")
                )

                # Analyse groupes
                group_columns = {}
                q_translated_final = None
                total = len(st.session_state.transcriptions)
                for gi, (name, bts) in enumerate(list(st.session_state.transcriptions.items()), start=1):
                    log(f"Analyse du Groupe {gi} — {name}")
                    text_tr = read_docx_text(bts)

                    # analyze_group doit :
                    # - traduire/normaliser les questions en target_lang (retour q_trans)
                    # - produire answers: dict index(str: "1","2",...) -> réponse résumée pro, dans la langue cible
                    q_trans, answers = analyze_group(
                        bridge, text_tr, questions, target_lang, guide_lang,
                        progress_cb=lambda p, base=gi - 1, tot=total: prog.progress(min(1.0, (base + p) / tot)),
                        log_cb=log,
                    )

                    # Première passe : verrouiller l'ordre final des questions traduites
                    if q_translated_final is None:
                        q_translated_final = q_trans

                    # Récupérer les réponses dans l'ordre des questions
                    ordered = []
                    for i in range(1, len(q_translated_final) + 1):
                        val = answers.get(str(i), "").strip()
                        ordered.append(val if val else "Non abordé")

                    group_columns[f"Groupe {gi}"] = ordered
                    log(f"  ✓ Groupe {gi} terminé.")

                if not q_translated_final:
                    st.error("Aucune question finale n'a été obtenue. Vérifiez le guide.")
                else:
                    # DataFrame final
                    df = pd.DataFrame({"Question": q_translated_final})
                    for gname, vals in group_columns.items():
                        # S'assurer que chaque colonne a la bonne longueur
                        if len(vals) < len(q_translated_final):
                            vals = vals + ["Non abordé"] * (len(q_translated_final) - len(vals))
                        df[gname] = vals

                    # Excel style pro
                    try:
                        xlsx_bytes = build_styled_excel(df)
                    except Exception as e:
                        log(f"Style Excel échoué ({e}), fallback sans style.")
                        buff = io.BytesIO()
                        with pd.ExcelWriter(buff, engine="openpyxl") as writer:
                            df.to_excel(writer, index=False, sheet_name="Analyse")
                        xlsx_bytes = buff.getvalue()

                    out_name = f"{excel_name}.xlsx"
                    st.session_state.generated_excels.append((out_name, xlsx_bytes))
                    st.success("Excel généré avec succès.")
                    prog.progress(1.0)
                    log("Excel généré.")

    if st.button("Fusionner des Excels"):
        st.session_state["merge_mode"] = True

    if st.session_state.get("merge_mode"):
        st.markdown("**Téléverser des Excels à fusionner (eux-mêmes produits par IPSOS ANALYZER)**")
        merge_files = st.file_uploader("Choisir des fichiers .xlsx (2+)", type=["xlsx"], accept_multiple_files=True, key="merge_upload")
        if st.button("Fusionner maintenant"):
            if not merge_files or len(merge_files) < 2:
                st.warning("Ajoutez au moins deux fichiers Excel.")
            else:
                dfs = [pd.read_excel(f) for f in merge_files]
                base = dfs[0]
                for d in dfs[1:]:
                    base = base.merge(d, on="Question", how="outer", suffixes=("", "_dup"))
                    to_drop = [c for c in base.columns if c.endswith("_dup")]
                    if to_drop:
                        base = base.drop(columns=to_drop)

                # Style rapide pour fusion
                try:
                    xlsx_bytes = (lambda df_merge: __import__('io').BytesIO() or None)  # dummy to keep scope
                except:
                    pass
                buff = io.BytesIO()
                with pd.ExcelWriter(buff, engine="openpyxl") as writer:
                    base.to_excel(writer, index=False, sheet_name="Fusion")
                st.session_state.generated_excels.append((f"{excel_name}_fusion.xlsx", buff.getvalue()))
                st.success("Fusion terminée.")
                log("Fichiers Excel fusionnés.")

# ----------------------------- JOURNAL -----------------------------
st.markdown("---")
st.markdown("#### 7) Journal")
st.text_area("Journal d'exécution", value="\n".join(st.session_state.log), height=260, key="log_area", label_visibility="collapsed")

# ----------------------------- DOWNLOADS & DESKTOP SAVE -----------------------------
st.markdown("---")
st.markdown("#### 8) Export")
if st.session_state.generated_excels:
    for nm, data in st.session_state.generated_excels[-3:][::-1]:
        st.download_button("Télécharger " + nm, data=data, file_name=nm,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    save_to_desktop = st.checkbox("Enregistrer aussi sur le Bureau (si autorisé par le système)", value=False)
    if save_to_desktop and st.button("Enregistrer maintenant sur le Bureau"):
        try:
            home = os.path.expanduser("~")
            desk = os.path.join(home, "Desktop")
            os.makedirs(desk, exist_ok=True)
            saved = []
            for nm, data in st.session_state.generated_excels:
                path = os.path.join(desk, nm)
                with open(path, "wb") as f:
                    f.write(data)
                saved.append(path)
            st.success("Fichiers enregistrés sur le Bureau :\n" + "\n".join(saved))
            log("Fichiers enregistrés sur le Bureau.")
        except Exception as e:
            st.error(f"Impossible d'enregistrer sur le Bureau : {e}")
            log(f"Echec d'enregistrement sur le Bureau : {e}")
else:
    st.info("Générez un Excel pour activer l'export.")

# ----------------------------- FOOTER -----------------------------
st.caption("© 2025 — IPSOS ANALYZER (prototype). Importez votre logo officiel dans assets/ pour remplacer l'espace réservé.")

